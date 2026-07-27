# coding=utf-8
"""exporter.py — Excel 导出器"""

import os
from typing import Dict, List
from datetime import datetime
from .models import UserConfig, StockHolding, AllocationResult
from .formatter import ResultFormatter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


HEADER_FONT = None
HEADER_FILL = None
HEADER_ALIGN = None


def _init_styles():
    global HEADER_FONT, HEADER_FILL, HEADER_ALIGN
    if HAS_OPENPYXL and HEADER_FONT is None:
        HEADER_FONT = Font(bold=True, size=11)
        HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
        HEADER_ALIGN = Alignment(horizontal="center")


class ExcelExporter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.fmt = ResultFormatter()

    def export(self, results: Dict[str, AllocationResult],
               config: UserConfig, holdings: List[StockHolding]) -> str:
        if not HAS_OPENPYXL:
            return self._export_csv_fallback(results, config, holdings)
        _init_styles()
        wb = Workbook()
        self._write_scenario_sheet(wb, results, config)
        self._write_detail_sheet(wb, results, holdings)
        self._write_readme_sheet(wb, config, holdings)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(self.output_path)
        return self.output_path

    def _export_csv_fallback(self, results, config, holdings):
        import csv
        base, _ = os.path.splitext(self.output_path)
        path = base + ".csv"
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["行情", "代码", "名称", "价格", "底仓", "活动仓", "总建议", "增仓", "范围"])
            for key, r in results.items():
                for d in r.stock_details:
                    w.writerow([r.scenario_name, d.stock_code, d.stock_name,
                                d.current_price, d.base_shares, d.active_shares,
                                d.total_shares, d.additional_shares, d.buy_range_str])
        return path

    def _write_scenario_sheet(self, wb, results, config):
        ws = wb.active
        ws.title = "场景对比总表"
        headers = ["项目", "行情下行", "行情震荡", "行情上行"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, HEADER_ALIGN
        keys = ["下行", "震荡", "上行"]
        rows = [
            ("资金总量", "C"), ("底仓金额", "B"), ("底仓占比", "BP"),
            ("活动仓金额", "A"), ("活动仓占比", "AP"),
            ("现金金额", "CASH"), ("现金占比", "CP"),
            ("买入强度", "I"), ("股票数量", "N"),
        ]
        for r_idx, (label, _) in enumerate(rows, 2):
            ws.cell(r_idx, 1, label).font = Font(bold=True) if HAS_OPENPYXL else None
        for r_idx, (label, kind) in enumerate(rows, 2):
            for c_idx, key in enumerate(keys, 2):
                r = results.get(key)
                if r is None: continue
                C = r.total_capital
                val = {"C": self.fmt.format_currency(C),
                       "B": self.fmt.format_currency(r.base_amount),
                       "BP": self.fmt.format_pct(r.base_pct),
                       "A": self.fmt.format_currency(r.active_amount),
                       "AP": self.fmt.format_pct(r.active_pct),
                       "CASH": self.fmt.format_currency(r.cash_amount),
                       "CP": self.fmt.format_pct(r.cash_pct),
                       "I": self.fmt.format_pct(r.buy_intensity),
                       "N": len(r.stock_details)}.get(kind, "")
                ws.cell(r_idx, c_idx, val)
        for c in range(1, 5):
            ws.column_dimensions[chr(64 + c)].width = 18

    def _write_detail_sheet(self, wb, results, holdings):
        ws = wb.create_sheet("个股数量范围明细")
        headers = ["股票代码", "股票名称", "当前价格", "板块", "风险等级", "已有持仓"]
        keys = ["下行", "震荡", "上行"]
        for k in keys:
            headers += [f"{k}_底仓", f"{k}_活动仓", f"{k}_总建议", f"{k}_增仓", f"{k}_范围"]
        headers.append("备注")
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font, cell.fill, cell.alignment = HEADER_FONT, HEADER_FILL, HEADER_ALIGN
        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="FFC7CE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        for r_idx, h in enumerate(holdings, 2):
            ws.cell(r_idx, 1, h.stock_code)
            ws.cell(r_idx, 2, h.stock_name)
            ws.cell(r_idx, 3, h.current_price).number_format = '#,##0.00'
            ws.cell(r_idx, 4, h.sector)
            ws.cell(r_idx, 5, h.risk_level)
            ws.cell(r_idx, 6, h.existing_shares)
            detail_for_code = {}
            for key in keys:
                r = results.get(key)
                if r:
                    d = next((x for x in r.stock_details if x.stock_code == h.stock_code), None)
                    if d: detail_for_code[key] = d
            for c_idx, key in enumerate(keys):
                d = detail_for_code.get(key)
                if d is None: continue
                col = 7 + c_idx * 5
                ws.cell(r_idx, col, d.base_shares)
                ws.cell(r_idx, col + 1, d.active_shares)
                ws.cell(r_idx, col + 2, d.total_shares)
                add_cell = ws.cell(r_idx, col + 3, d.additional_shares)
                if d.additional_shares < 0: add_cell.fill = red
                elif d.additional_shares == 0: add_cell.fill = green
                else: add_cell.fill = yellow
                ws.cell(r_idx, col + 4, d.buy_range_str)
            last_d = list(detail_for_code.values())[-1] if detail_for_code else None
            ws.cell(r_idx, len(headers), last_d.note if last_d else "")

    def _write_readme_sheet(self, wb, config, holdings):
        ws = wb.create_sheet("参数与说明")
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 55
        lines = [
            ("输入参数", ""), ("资金总量", self.fmt.format_currency(config.total_capital)),
            ("底仓比例", self.fmt.format_pct(config.base_position_ratio)),
            ("活动仓比例", self.fmt.format_pct(config.active_position_ratio)),
            ("现金比例", self.fmt.format_pct(config.cash_reserve_ratio)),
            ("股票数量", len(holdings)), ("分配策略", config.allocation_strategy),
            ("", ""), ("计算公式", ""),
            ("底仓金额", "C × b × (1 + 场景调整)"),
            ("活动仓金额", "C × a × (1 + 场景调整)"),
            ("现金金额", "MAX(C × c × (1 + 场景调整), C × 20%)"),
            ("", ""), ("风险提示", ""),
            ("", "本方案仅为参考, 不构成投资建议"),
            ("", "现金底线 20% 为硬性约束"),
            ("", f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
        ]
        for r, (a, b) in enumerate(lines, 1):
            if a:
                ws.cell(r, 1, a).font = Font(bold=True) if HAS_OPENPYXL else None
            ws.cell(r, 2, b)
